"""
城治学生会物资管理系统 - 后端主程序（升级版）
功能：用户登录、物资管理、数据大屏、扫码领用、扫码归还、图片上传、物资查询
"""

import os
import time
import sqlite3
import uuid
import random
import string
import smtplib
from urllib.parse import quote
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from datetime import datetime, timedelta
from typing import Optional, List

# ==================== 版本信息 ====================
VERSION = "1.9.2"
VERSION_DATE = "2026-08-30"

# 加载 .env 文件（纯 Python 实现，不依赖 python-dotenv）
def load_env():
    env_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), ".env")
    if not os.path.exists(env_path):
        return
    with open(env_path, "r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue
            key, _, value = line.partition("=")
            key = key.strip()
            value = value.strip().strip('"').strip("'")
            if key and key not in os.environ:
                os.environ[key] = value

load_env()

from fastapi import FastAPI, Depends, HTTPException, status, Header, UploadFile, File, Form
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse, StreamingResponse, Response
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
from jose import JWTError, jwt
import bcrypt
import openpyxl
from io import BytesIO

# ==================== 配置 ====================
DATABASE_FILE = os.environ.get("DATABASE_FILE", "data/material.db")
UPLOAD_DIR = os.environ.get("UPLOAD_DIR", "static/uploads")
SECRET_KEY = os.environ.get("SECRET_KEY", "material-system-secret-key-change-me-in-production")
ALGORITHM = "HS256"
ACCESS_TOKEN_EXPIRE_HOURS = 24 * 7

# CORS 允许的来源，多个用逗号分隔，默认全部允许（开发用）
CORS_ORIGINS = os.environ.get("CORS_ORIGINS", "*")

# ====== 邮箱SMTP配置（忘记密码功能需要配置）======
# 配置说明：通过环境变量设置，不要把密码写在代码里
# 1. QQ邮箱：SMTP_SERVER="smtp.qq.com", SMTP_PORT=465, SMTP_USER="你的QQ邮箱@qq.com", SMTP_PASSWORD="QQ邮箱授权码"
#    授权码获取：QQ邮箱→设置→账户→POP3/SMTP服务→开启→生成授权码
# 2. 163邮箱：SMTP_SERVER="smtp.163.com", SMTP_PORT=465, SMTP_USER="你的邮箱@163.com", SMTP_PASSWORD="授权码"
# 3. 不配置则忘记密码功能不可用，不影响其他功能
SMTP_SERVER = os.environ.get("SMTP_SERVER", "")
SMTP_PORT = int(os.environ.get("SMTP_PORT", "465"))
SMTP_USER = os.environ.get("SMTP_USER", "")
SMTP_PASSWORD = os.environ.get("SMTP_PASSWORD", "")
SMTP_ENABLED = bool(SMTP_SERVER and SMTP_USER and SMTP_PASSWORD)

# 验证码存储（内存）：{email: {"code": "123456", "expire": datetime}}
reset_codes = {}
register_codes = {}

# 确保上传目录存在
os.makedirs(UPLOAD_DIR, exist_ok=True)

# ==================== 初始化 ====================
app = FastAPI(title="城治学生会物资管理系统")

app.add_middleware(
    CORSMiddleware,
    allow_origins=CORS_ORIGINS.split(",") if CORS_ORIGINS != "*" else ["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# 密码加密
def hash_password(password: str) -> str:
    return bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')

def verify_password(plain_password: str, hashed_password: str) -> bool:
    try:
        return bcrypt.checkpw(plain_password.encode('utf-8'), hashed_password.encode('utf-8'))
    except Exception:
        return False

# ==================== 数据库初始化 ====================
def init_db():
    # 确保数据库目录存在
    db_dir = os.path.dirname(DATABASE_FILE)
    if db_dir:
        os.makedirs(db_dir, exist_ok=True)
    conn = sqlite3.connect(DATABASE_FILE)
    conn.row_factory = sqlite3.Row
    c = conn.cursor()

    # 用户表
    c.execute("""
        CREATE TABLE IF NOT EXISTS users (
            id TEXT PRIMARY KEY,
            username TEXT UNIQUE NOT NULL,
            password_hash TEXT NOT NULL,
            real_name TEXT NOT NULL,
            phone TEXT DEFAULT '',
            email TEXT DEFAULT '',
            department TEXT DEFAULT '',
            role TEXT DEFAULT 'user',
            created_at TEXT DEFAULT CURRENT_TIMESTAMP
        )
    """)

    # 物资表
    c.execute("""
        CREATE TABLE IF NOT EXISTS materials (
            id TEXT PRIMARY KEY,
            name TEXT NOT NULL,
            qr_code TEXT UNIQUE NOT NULL,
            spec TEXT,
            unit TEXT DEFAULT '个',
            total_stock INTEGER DEFAULT 0,
            available_stock INTEGER DEFAULT 0,
            location TEXT,
            image TEXT DEFAULT '',
            operator TEXT DEFAULT '',
            created_at TEXT DEFAULT CURRENT_TIMESTAMP
        )
    """)

    # 物资位置表（同一物资可存放在多个位置，每个位置独立库存）
    c.execute("""
        CREATE TABLE IF NOT EXISTS material_locations (
            id TEXT PRIMARY KEY,
            material_id TEXT NOT NULL,
            location TEXT NOT NULL,
            stock INTEGER DEFAULT 0,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP,
            UNIQUE(material_id, location)
        )
    """)

    # 迁移现有物资的位置数据到位置表
    c.execute("SELECT id, location, total_stock FROM materials WHERE location IS NOT NULL AND location != ''")
    existing_materials = c.fetchall()
    for m in existing_materials:
        # 检查是否已经迁移过
        c.execute("SELECT COUNT(*) FROM material_locations WHERE material_id=?", (m["id"],))
        if c.fetchone()[0] == 0:
            loc_id = str(uuid.uuid4())
            c.execute("INSERT INTO material_locations (id, material_id, location, stock) VALUES (?, ?, ?, ?)",
                      (loc_id, m["id"], m["location"], m["total_stock"]))
    conn.commit()

    # 领用归还表
    c.execute("""
        CREATE TABLE IF NOT EXISTS transactions (
            id TEXT PRIMARY KEY,
            material_id TEXT NOT NULL,
            user_id TEXT NOT NULL,
            type TEXT NOT NULL,
            quantity INTEGER NOT NULL,
            status TEXT DEFAULT 'active',
            activity_name TEXT DEFAULT '',
            phone TEXT DEFAULT '',
            borrow_image TEXT DEFAULT '',
            return_time TEXT DEFAULT '',
            return_location TEXT DEFAULT '',
            return_image TEXT DEFAULT '',
            borrowed_at TEXT DEFAULT CURRENT_TIMESTAMP,
            returned_at TEXT,
            remark TEXT
        )
    """)

    # 索引
    c.execute("CREATE INDEX IF NOT EXISTS idx_materials_qr ON materials(qr_code)")
    c.execute("CREATE INDEX IF NOT EXISTS idx_transactions_material ON transactions(material_id)")
    c.execute("CREATE INDEX IF NOT EXISTS idx_transactions_user ON transactions(user_id)")

    # 升级旧数据库：增加新字段（如果不存在）
    def add_column_if_not_exists(table, column, definition):
        c.execute(f"PRAGMA table_info({table})")
        columns = [row[1] for row in c.fetchall()]
        if column not in columns:
            c.execute(f"ALTER TABLE {table} ADD COLUMN {column} {definition}")

    add_column_if_not_exists("materials", "image", "TEXT DEFAULT ''")
    add_column_if_not_exists("materials", "operator", "TEXT DEFAULT ''")
    add_column_if_not_exists("transactions", "activity_name", "TEXT DEFAULT ''")
    add_column_if_not_exists("transactions", "phone", "TEXT DEFAULT ''")
    add_column_if_not_exists("transactions", "borrow_image", "TEXT DEFAULT ''")
    add_column_if_not_exists("transactions", "return_time", "TEXT DEFAULT ''")
    add_column_if_not_exists("transactions", "return_location", "TEXT DEFAULT ''")
    add_column_if_not_exists("transactions", "return_image", "TEXT DEFAULT ''")
    add_column_if_not_exists("transactions", "borrow_location", "TEXT DEFAULT ''")
    add_column_if_not_exists("users", "phone", "TEXT DEFAULT ''")
    add_column_if_not_exists("users", "email", "TEXT DEFAULT ''")
    add_column_if_not_exists("users", "department", "TEXT DEFAULT ''")

    # 初始化管理员
    c.execute("SELECT COUNT(*) FROM users WHERE username = 'admin'")
    if c.fetchone()[0] == 0:
        c.execute("""
            INSERT INTO users (id, username, password_hash, real_name, phone, role)
            VALUES (?, ?, ?, ?, ?, 'admin')
        """, (str(uuid.uuid4()), "admin", hash_password("admin123"), "系统管理员", ""))

    # 初始化测试用户（中文用户名，密码123456）
    test_users = [
        ("张三", "13800138001"),
        ("李四", "13800138002"),
        ("王五", "13800138003"),
    ]
    for name, phone in test_users:
        c.execute("SELECT COUNT(*) FROM users WHERE username = ?", (name,))
        if c.fetchone()[0] == 0:
            c.execute("""
                INSERT INTO users (id, username, password_hash, real_name, phone, role)
                VALUES (?, ?, ?, ?, ?, 'user')
            """, (str(uuid.uuid4()), name, hash_password("123456"), name, phone))

    # 初始化测试物资
    c.execute("SELECT COUNT(*) FROM materials")
    if c.fetchone()[0] == 0:
        samples = [
            (str(uuid.uuid4()), "笔记本电脑", "MAT-0001", "ThinkPad X1", "台", 5, 5, "办公室A柜", ""),
            (str(uuid.uuid4()), "投影仪", "MAT-0002", "爱普生 CB-X06", "台", 2, 2, "会议室B柜", ""),
            (str(uuid.uuid4()), "万用表", "MAT-0003", "数字式", "个", 10, 10, "工具柜C层", ""),
            (str(uuid.uuid4()), "马克笔", "MAT-0004", "黑色", "支", 50, 50, "文具柜", ""),
            (str(uuid.uuid4()), "订书机", "MAT-0005", "标准型", "个", 8, 8, "文具柜", ""),
        ]
        c.executemany("""
            INSERT INTO materials (id, name, qr_code, spec, unit, total_stock, available_stock, location, image)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, samples)

    # 确保 admin 绑定邮箱（用于忘记密码）
    if SMTP_ENABLED:
        c.execute("UPDATE users SET email = ? WHERE username = 'admin'", (SMTP_USER,))

    conn.commit()
    conn.close()
    print("  [数据库] 本地数据库初始化完成（升级版）")

init_db()

# ==================== 数据库连接 ====================
def get_db():
    conn = sqlite3.connect(DATABASE_FILE)
    conn.row_factory = sqlite3.Row
    try:
        yield conn
    finally:
        conn.close()

# ==================== 数据模型 ====================
class LoginRequest(BaseModel):
    username: str
    password: str

class ChangePasswordRequest(BaseModel):
    username: str
    old_password: str
    new_password: str

class RegisterRequest(BaseModel):
    username: str
    password: str = ""
    real_name: str = ""
    phone: str = ""
    email: str = ""
    department: str = ""
    code: str = ""

class UserCreateRequest(BaseModel):
    username: str
    password: str = "123456"
    real_name: str = ""
    phone: str = ""
    email: str = ""
    department: str = ""
    role: str = "user"

class UserUpdateRequest(BaseModel):
    username: Optional[str] = None
    real_name: Optional[str] = None
    phone: Optional[str] = None
    email: Optional[str] = None
    department: Optional[str] = None
    role: Optional[str] = None
    password: Optional[str] = None

class MaterialCreate(BaseModel):
    name: str
    spec: Optional[str] = ""
    unit: str = "个"
    total_stock: int = 0
    location: Optional[str] = ""
    image: Optional[str] = ""

class MaterialUpdate(BaseModel):
    name: Optional[str] = None
    spec: Optional[str] = None
    unit: Optional[str] = None
    total_stock: Optional[int] = None
    location: Optional[str] = None
    image: Optional[str] = None

class BorrowRequest(BaseModel):
    material_id: str
    quantity: int = 1
    activity_name: str = ""
    phone: str = ""
    borrow_image: str = ""
    location: str = ""

class ReturnRequest(BaseModel):
    material_id: str
    quantity: int = 1
    return_time: str = ""
    return_location: str = ""
    return_image: str = ""

# 公开接口请求模型（带登录信息，无需预先登录）
class PublicBorrowRequest(BaseModel):
    username: str
    password: str
    material_id: str
    quantity: int = 1
    activity_name: str = ""
    phone: str = ""
    borrow_image: str = ""
    location: str = ""

class PublicReturnRequest(BaseModel):
    username: str
    password: str
    material_id: str
    quantity: int = 1
    return_time: str = ""
    return_location: str = ""
    return_image: str = ""

# ==================== 认证 ====================
def create_access_token(data: dict) -> str:
    to_encode = data.copy()
    expire = datetime.utcnow() + timedelta(hours=ACCESS_TOKEN_EXPIRE_HOURS)
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, SECRET_KEY, algorithm=ALGORITHM)

def get_current_user(authorization: str = Header(...), conn=Depends(get_db)):
    credentials_exception = HTTPException(
        status_code=status.HTTP_401_UNAUTHORIZED,
        detail="登录已过期，请重新登录",
        headers={"WWW-Authenticate": "Bearer"},
    )
    try:
        if not authorization.startswith("Bearer "):
            raise credentials_exception
        token = authorization[7:]
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        user_id: str = payload.get("sub")
        if user_id is None:
            raise credentials_exception
    except JWTError:
        raise credentials_exception

    cur = conn.cursor()
    cur.execute("SELECT id, username, real_name, phone, email, department, role FROM users WHERE id = ?", (user_id,))
    user = cur.fetchone()
    if user is None:
        raise credentials_exception
    return dict(user)

# ==================== API 接口 ====================

# ==================== 版本信息 ====================
@app.get("/api/version")
def get_version():
    changelog = []
    try:
        changelog_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "CHANGELOG.md")
        with open(changelog_path, "r", encoding="utf-8") as f:
            content = f.read()
        # 简单解析，返回所有版本
        import re
        versions = re.findall(r'## \[([^\]]+)\] - ([^\n]+)\n(.*?)(?=\n## \[|\Z)', content, re.DOTALL)
        for v, date, body in versions:
            changelog.append({"version": v, "date": date.strip(), "content": body.strip()})
    except Exception:
        pass
    return {
        "version": VERSION,
        "date": VERSION_DATE,
        "changelog": changelog
    }

@app.post("/api/login")
def login(req: LoginRequest, conn=Depends(get_db)):
    cur = conn.cursor()
    cur.execute("SELECT * FROM users WHERE username = ?", (req.username,))
    user = cur.fetchone()
    if not user or not verify_password(req.password, user["password_hash"]):
        raise HTTPException(status_code=400, detail="账号或密码错误")
    access_token = create_access_token(data={"sub": user["id"]})
    return {
        "token": access_token,
        "user_id": user["id"],
        "username": user["username"],
        "real_name": user["real_name"],
        "phone": user["phone"],
        "email": user["email"],
        "department": user["department"],
        "role": user["role"],
    }

# ---------- 修改密码 ----------
@app.post("/api/change-password")
def change_password(req: ChangePasswordRequest, conn=Depends(get_db)):
    if len(req.new_password) < 4:
        raise HTTPException(status_code=400, detail="新密码至少4位")
    cur = conn.cursor()
    cur.execute("SELECT * FROM users WHERE username = ?", (req.username,))
    user = cur.fetchone()
    if not user:
        raise HTTPException(status_code=400, detail="账号不存在")
    if not verify_password(req.old_password, user["password_hash"]):
        raise HTTPException(status_code=400, detail="原密码错误")
    new_hash = hash_password(req.new_password)
    cur.execute("UPDATE users SET password_hash = ? WHERE id = ?", (new_hash, user["id"]))
    conn.commit()
    return {"success": True, "message": "密码修改成功，请用新密码登录"}

# ---------- 用户注册 ----------
@app.post("/api/register")
def register(req: RegisterRequest, conn=Depends(get_db)):
    username = req.username.strip()
    if not username:
        raise HTTPException(status_code=400, detail="用户名不能为空")
    if len(username) > 20:
        raise HTTPException(status_code=400, detail="用户名不能超过20个字符")
    email = req.email.strip().lower()
    if not email:
        raise HTTPException(status_code=400, detail="请输入邮箱")
    if not req.code:
        raise HTTPException(status_code=400, detail="请输入邮箱验证码")
    # 验证验证码
    record = register_codes.get(email)
    if not record:
        raise HTTPException(status_code=400, detail="请先获取邮箱验证码")
    if datetime.now() > record["expire"]:
        del register_codes[email]
        raise HTTPException(status_code=400, detail="验证码已过期，请重新获取")
    if record["code"] != req.code.strip():
        raise HTTPException(status_code=400, detail="验证码错误")
    cur = conn.cursor()
    cur.execute("SELECT id FROM users WHERE username = ?", (username,))
    if cur.fetchone():
        raise HTTPException(status_code=400, detail="该用户名已被注册")
    cur.execute("SELECT id FROM users WHERE email = ?", (email,))
    if cur.fetchone():
        raise HTTPException(status_code=400, detail="该邮箱已被注册")
    password = req.password if req.password else "123456"
    if len(password) < 4:
        raise HTTPException(status_code=400, detail="密码至少4位")
    real_name = req.real_name if req.real_name else username
    user_id = str(uuid.uuid4())
    cur.execute("""
        INSERT INTO users (id, username, password_hash, real_name, phone, email, department, role)
        VALUES (?, ?, ?, ?, ?, ?, ?, 'user')
    """, (user_id, username, hash_password(password), real_name, req.phone, email, req.department))
    conn.commit()
    # 删除已使用的验证码
    if email in register_codes:
        del register_codes[email]
    return {"success": True, "message": "注册成功，请登录", "username": username}

# ---------- 用户管理（管理员）----------
@app.get("/api/users")
def list_users(conn=Depends(get_db), user=Depends(get_current_user)):
    if user["role"] != "admin":
        raise HTTPException(status_code=403, detail="只有管理员可以查看用户")
    cur = conn.cursor()
    cur.execute("""
        SELECT id, username, real_name, phone, email, department, role, created_at
        FROM users ORDER BY created_at DESC
    """)
    return [dict(u) for u in cur.fetchall()]

@app.post("/api/users")
def create_user(req: UserCreateRequest, conn=Depends(get_db), user=Depends(get_current_user)):
    if user["role"] != "admin":
        raise HTTPException(status_code=403, detail="只有管理员可以创建用户")
    username = req.username.strip()
    if not username:
        raise HTTPException(status_code=400, detail="用户名不能为空")
    cur = conn.cursor()
    cur.execute("SELECT id FROM users WHERE username = ?", (username,))
    if cur.fetchone():
        raise HTTPException(status_code=400, detail="该用户名已存在")
    password = req.password if req.password else "123456"
    real_name = req.real_name if req.real_name else username
    user_id = str(uuid.uuid4())
    cur.execute("""
        INSERT INTO users (id, username, password_hash, real_name, phone, email, department, role)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?)
    """, (user_id, username, hash_password(password), real_name, req.phone, req.email, req.department, req.role))
    conn.commit()
    return {"success": True, "message": "用户创建成功", "user_id": user_id}

@app.put("/api/users/{user_id}")
def update_user(user_id: str, req: UserUpdateRequest, conn=Depends(get_db), user=Depends(get_current_user)):
    if user["role"] != "admin":
        raise HTTPException(status_code=403, detail="只有管理员可以修改用户")
    cur = conn.cursor()
    cur.execute("SELECT id FROM users WHERE id = ?", (user_id,))
    if not cur.fetchone():
        raise HTTPException(status_code=404, detail="用户不存在")
    updates = []
    params = []
    if req.username is not None:
        new_username = req.username.strip()
        if not new_username:
            raise HTTPException(status_code=400, detail="用户名不能为空")
        cur.execute("SELECT id FROM users WHERE username = ? AND id != ?", (new_username, user_id))
        if cur.fetchone():
            raise HTTPException(status_code=400, detail="该用户名已被使用")
        updates.append("username = ?")
        params.append(new_username)
    if req.real_name is not None:
        updates.append("real_name = ?")
        params.append(req.real_name)
    if req.phone is not None:
        updates.append("phone = ?")
        params.append(req.phone)
    if req.email is not None:
        updates.append("email = ?")
        params.append(req.email)
    if req.department is not None:
        updates.append("department = ?")
        params.append(req.department)
    if req.role is not None:
        if req.role not in ("admin", "user"):
            raise HTTPException(status_code=400, detail="角色只能是 admin 或 user")
        updates.append("role = ?")
        params.append(req.role)
    if req.password is not None:
        if len(req.password) < 4:
            raise HTTPException(status_code=400, detail="密码至少4位")
        updates.append("password_hash = ?")
        params.append(hash_password(req.password))
    if updates:
        params.append(user_id)
        cur.execute(f"UPDATE users SET {', '.join(updates)} WHERE id = ?", params)
        conn.commit()
    return {"success": True, "message": "用户信息已更新"}

# ---------- 修改自己的个人信息（任何登录用户）----------
@app.put("/api/profile")
def update_profile(req: UserUpdateRequest, conn=Depends(get_db), user=Depends(get_current_user)):
    cur = conn.cursor()
    updates = []
    params = []
    if req.username is not None:
        new_username = req.username.strip()
        if not new_username:
            raise HTTPException(status_code=400, detail="用户名不能为空")
        cur.execute("SELECT id FROM users WHERE username = ? AND id != ?", (new_username, user["id"]))
        if cur.fetchone():
            raise HTTPException(status_code=400, detail="该用户名已被使用")
        updates.append("username = ?")
        params.append(new_username)
    if req.real_name is not None:
        updates.append("real_name = ?")
        params.append(req.real_name)
    if req.phone is not None:
        updates.append("phone = ?")
        params.append(req.phone)
    if req.email is not None:
        updates.append("email = ?")
        params.append(req.email)
    if req.department is not None:
        updates.append("department = ?")
        params.append(req.department)
    if req.password is not None:
        if len(req.password) < 4:
            raise HTTPException(status_code=400, detail="密码至少4位")
        updates.append("password_hash = ?")
        params.append(hash_password(req.password))
    if updates:
        params.append(user["id"])
        cur.execute(f"UPDATE users SET {', '.join(updates)} WHERE id = ?", params)
        conn.commit()
        # 更新localStorage中的用户信息
        cur.execute("SELECT id, username, real_name, phone, email, department, role FROM users WHERE id = ?", (user["id"],))
        updated = cur.fetchone()
        return {"success": True, "message": "个人信息已更新", "user": dict(updated)}
    return {"success": True, "message": "没有需要更新的内容"}

@app.delete("/api/users/{user_id}")
def delete_user(user_id: str, conn=Depends(get_db), user=Depends(get_current_user)):
    if user["role"] != "admin":
        raise HTTPException(status_code=403, detail="只有管理员可以删除用户")
    if user_id == user["id"]:
        raise HTTPException(status_code=400, detail="不能删除当前登录的账号")
    cur = conn.cursor()
    cur.execute("SELECT username FROM users WHERE id = ?", (user_id,))
    target = cur.fetchone()
    if not target:
        raise HTTPException(status_code=404, detail="用户不存在")
    cur.execute("DELETE FROM users WHERE id = ?", (user_id,))
    cur.execute("DELETE FROM transactions WHERE user_id = ?", (user_id,))
    conn.commit()
    return {"success": True, "message": f"用户 {target['username']} 已删除"}

# ---------- 批量删除用户 ----------
@app.post("/api/users/batch-delete")
def batch_delete_users(req: dict, conn=Depends(get_db), user=Depends(get_current_user)):
    if user["role"] != "admin":
        raise HTTPException(status_code=403, detail="只有管理员可以删除用户")
    ids = req.get("ids", [])
    if not ids:
        raise HTTPException(status_code=400, detail="请选择要删除的用户")
    if user["id"] in ids:
        raise HTTPException(status_code=400, detail="不能删除当前登录的账号")
    cur = conn.cursor()
    deleted = 0
    for uid in ids:
        cur.execute("SELECT id FROM users WHERE id=?", (uid,))
        if cur.fetchone():
            cur.execute("DELETE FROM users WHERE id=?", (uid,))
            cur.execute("DELETE FROM transactions WHERE user_id=?", (uid,))
            deleted += 1
    conn.commit()
    return {"success": True, "message": f"已删除 {deleted} 个用户"}

# ---------- 下载用户导入模板 ----------
@app.get("/api/users/template")
def download_user_template(user=Depends(get_current_user)):
    if user["role"] != "admin":
        raise HTTPException(status_code=403, detail="只有管理员可以下载模板")
    output = BytesIO()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "用户导入模板"
    # 表头
    headers = ["用户名*", "真实姓名", "部门", "电话", "邮箱", "初始密码"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = openpyxl.styles.Font(bold=True, color="FFFFFF")
        cell.fill = openpyxl.styles.PatternFill(start_color="667EEA", end_color="667EEA", fill_type="solid")
    # 示例数据
    examples = [
        ["zhangsan", "张三", "办公室", "13800138000", "zhangsan@example.com", "123456"],
        ["lisi", "李四", "学习部", "13900139000", "lisi@example.com", "123456"],
    ]
    for row_idx, row_data in enumerate(examples, 2):
        for col_idx, val in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=val)
    # 填写说明
    ws.cell(row=5, column=1, value="填写说明：").font = openpyxl.styles.Font(bold=True, color="FF0000")
    notes = [
        "1. 用户名必填，不能重复，支持中文或英文",
        "2. 真实姓名不填则与用户名相同",
        "3. 部门可选：办公室/学习部/体育部/文娱部/志工部/宣传部/马列部/组织部/权益部/社团部/主席团/副书记",
        "4. 初始密码不填则默认123456",
        "5. 邮箱用于忘记密码找回，建议填写",
        "6. 带*为必填项",
    ]
    for i, note in enumerate(notes, 6):
        ws.cell(row=i, column=1, value=note)
    # 列宽
    widths = [15, 12, 12, 15, 25, 12]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    wb.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=user_template.xlsx"}
    )

# ---------- 批量导入用户 ----------
@app.post("/api/users/import")
def import_users(file: UploadFile = File(...), conn=Depends(get_db), user=Depends(get_current_user)):
    if user["role"] != "admin":
        raise HTTPException(status_code=403, detail="只有管理员可以导入用户")
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="请上传 .xlsx 格式的 Excel 文件")
    try:
        content = file.file.read()
        wb = openpyxl.load_workbook(BytesIO(content))
        ws = wb.active
        cur = conn.cursor()
        imported = 0
        skipped = []
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
            if row_idx > 4:  # 跳过示例和说明行
                break
            username = str(row[0]).strip() if row[0] else ""
            if not username or username.startswith("填写说明") or username.startswith("1."):
                continue
            # 检查用户名是否已存在
            cur.execute("SELECT id FROM users WHERE username=?", (username,))
            if cur.fetchone():
                skipped.append(f"第{row_idx}行：{username}（已存在）")
                continue
            real_name = str(row[1]).strip() if row[1] else username
            department = str(row[2]).strip() if row[2] else ""
            phone = str(row[3]).strip() if row[3] else ""
            email = str(row[4]).strip() if row[4] else ""
            password = str(row[5]).strip() if row[5] else "123456"
            user_id = str(uuid.uuid4())
            cur.execute("""
                INSERT INTO users (id, username, password_hash, real_name, phone, email, department, role)
                VALUES (?, ?, ?, ?, ?, ?, ?, 'user')
            """, (user_id, username, hash_password(password), real_name, phone, email, department))
            imported += 1
        # 继续读取后面的行（如果有更多数据）
        for row_idx, row in enumerate(ws.iter_rows(min_row=5, values_only=True), 5):
            username = str(row[0]).strip() if row[0] else ""
            if not username or username.startswith("填写说明") or username.startswith("1.") or username.startswith("2.") or username.startswith("3.") or username.startswith("4.") or username.startswith("5.") or username.startswith("6."):
                continue
            cur.execute("SELECT id FROM users WHERE username=?", (username,))
            if cur.fetchone():
                skipped.append(f"第{row_idx}行：{username}（已存在）")
                continue
            real_name = str(row[1]).strip() if row[1] else username
            department = str(row[2]).strip() if row[2] else ""
            phone = str(row[3]).strip() if row[3] else ""
            email = str(row[4]).strip() if row[4] else ""
            password = str(row[5]).strip() if row[5] else "123456"
            user_id = str(uuid.uuid4())
            cur.execute("""
                INSERT INTO users (id, username, password_hash, real_name, phone, email, department, role)
                VALUES (?, ?, ?, ?, ?, ?, ?, 'user')
            """, (user_id, username, hash_password(password), real_name, phone, email, department))
            imported += 1
        conn.commit()
        return {"success": True, "message": f"成功导入 {imported} 个用户", "imported": imported, "skipped": skipped}
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=f"导入失败：{str(e)}")

# ---------- 导出用户 ----------
@app.get("/api/users/export")
def export_users(conn=Depends(get_db), user=Depends(get_current_user)):
    if user["role"] != "admin":
        raise HTTPException(status_code=403, detail="只有管理员可以导出用户")
    cur = conn.cursor()
    cur.execute("""
        SELECT username, real_name, department, phone, email, role, created_at
        FROM users ORDER BY created_at DESC
    """)
    users = cur.fetchall()
    output = BytesIO()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "用户列表"
    headers = ["用户名", "真实姓名", "部门", "电话", "邮箱", "角色", "注册时间"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = openpyxl.styles.Font(bold=True, color="FFFFFF")
        cell.fill = openpyxl.styles.PatternFill(start_color="667EEA", end_color="667EEA", fill_type="solid")
    for row_idx, u in enumerate(users, 2):
        ws.cell(row=row_idx, column=1, value=u["username"])
        ws.cell(row=row_idx, column=2, value=u["real_name"])
        ws.cell(row=row_idx, column=3, value=u["department"] or "")
        ws.cell(row=row_idx, column=4, value=u["phone"] or "")
        ws.cell(row=row_idx, column=5, value=u["email"] or "")
        ws.cell(row=row_idx, column=6, value="管理员" if u["role"] == "admin" else "普通用户")
        ws.cell(row=row_idx, column=7, value=u["created_at"] or "")
    widths = [15, 12, 12, 15, 25, 10, 20]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    wb.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=users_export.xlsx"}
    )

# ==================== 忘记密码（邮箱验证码）====================
class SendCodeRequest(BaseModel):
    email: str

class ResetPasswordRequest(BaseModel):
    email: str
    code: str
    new_password: str

def send_email(to_email: str, subject: str, content: str) -> bool:
    """发送邮件，返回是否成功"""
    if not SMTP_ENABLED:
        return False
    try:
        msg = MIMEMultipart()
        msg["From"] = SMTP_USER
        msg["To"] = to_email
        msg["Subject"] = subject
        msg.attach(MIMEText(content, "plain", "utf-8"))
        with smtplib.SMTP_SSL(SMTP_SERVER, SMTP_PORT) as server:
            server.login(SMTP_USER, SMTP_PASSWORD)
            server.sendmail(SMTP_USER, to_email, msg.as_string())
        return True
    except Exception as e:
        print(f"[邮件发送失败] {e}")
        return False

@app.post("/api/send-register-code")
def send_register_code(req: SendCodeRequest, conn=Depends(get_db)):
    if not SMTP_ENABLED:
        raise HTTPException(status_code=400, detail="系统未配置邮箱，注册功能暂不可用，请联系管理员")
    email = req.email.strip().lower()
    if not email:
        raise HTTPException(status_code=400, detail="请输入邮箱")
    # 检查邮箱是否已被注册
    cur = conn.cursor()
    cur.execute("SELECT id FROM users WHERE email = ?", (email,))
    if cur.fetchone():
        raise HTTPException(status_code=400, detail="该邮箱已被注册")
    # 生成6位验证码
    code = ''.join(random.choices(string.digits, k=6))
    register_codes[email] = {"code": code, "expire": datetime.now() + timedelta(minutes=10)}
    # 发送邮件
    subject = "城治学生会物资管理系统 - 注册验证码"
    content = f"""您好！

您正在注册城治学生会物资管理系统账号。

邮箱：{email}

验证码：{code}

该验证码10分钟内有效，请勿泄露给他人。
如非本人操作，请忽略此邮件。

—— 城治学生会物资管理系统"""
    success = send_email(email, subject, content)
    if not success:
        raise HTTPException(status_code=500, detail="验证码发送失败，请检查邮箱配置或稍后重试")
    return {"success": True, "message": f"验证码已发送到 {email}，请注意查收（10分钟内有效）"}

@app.post("/api/send-reset-code")
def send_reset_code(req: SendCodeRequest, conn=Depends(get_db)):
    if not SMTP_ENABLED:
        raise HTTPException(status_code=400, detail="系统未配置邮箱，忘记密码功能暂不可用，请联系管理员")
    email = req.email.strip().lower()
    if not email:
        raise HTTPException(status_code=400, detail="请输入邮箱")
    # 检查邮箱是否绑定了账号
    cur = conn.cursor()
    cur.execute("SELECT id, real_name, username, department FROM users WHERE email = ?", (email,))
    user = cur.fetchone()
    if not user:
        raise HTTPException(status_code=400, detail="该邮箱未绑定任何账号")
    # 生成6位验证码
    code = ''.join(random.choices(string.digits, k=6))
    reset_codes[email] = {"code": code, "expire": datetime.now() + timedelta(minutes=10)}
    # 发送邮件
    dept = user['department'] if user['department'] else '（未填写）'
    subject = "城治学生会物资管理系统 - 密码重置验证码"
    content = f"""您好，{user['real_name']}！

您正在重置城治学生会物资管理系统的登录密码。

账号：{user['username']}
部门：{dept}

验证码：{code}

该验证码10分钟内有效，请勿泄露给他人。
如非本人操作，请忽略此邮件。

—— 城治学生会物资管理系统"""
    success = send_email(email, subject, content)
    if not success:
        raise HTTPException(status_code=500, detail="验证码发送失败，请检查邮箱配置或稍后重试")
    return {"success": True, "message": f"验证码已发送到 {email}，请注意查收（10分钟内有效）"}

@app.post("/api/reset-password")
def reset_password(req: ResetPasswordRequest, conn=Depends(get_db)):
    email = req.email.strip().lower()
    code = req.code.strip()
    new_password = req.new_password
    if len(new_password) < 4:
        raise HTTPException(status_code=400, detail="新密码至少4位")
    # 验证验证码
    record = reset_codes.get(email)
    if not record:
        raise HTTPException(status_code=400, detail="请先获取验证码")
    if datetime.now() > record["expire"]:
        del reset_codes[email]
        raise HTTPException(status_code=400, detail="验证码已过期，请重新获取")
    if record["code"] != code:
        raise HTTPException(status_code=400, detail="验证码错误")
    # 重置密码
    cur = conn.cursor()
    cur.execute("SELECT id FROM users WHERE email = ?", (email,))
    user = cur.fetchone()
    if not user:
        raise HTTPException(status_code=400, detail="该邮箱未绑定任何账号")
    new_hash = hash_password(new_password)
    cur.execute("UPDATE users SET password_hash = ? WHERE id = ?", (new_hash, user["id"]))
    conn.commit()
    # 删除已使用的验证码
    del reset_codes[email]
    return {"success": True, "message": "密码重置成功，请用新密码登录"}

@app.get("/api/email-config-status")
def email_config_status():
    return {"enabled": SMTP_ENABLED, "message": "邮箱已配置" if SMTP_ENABLED else "邮箱未配置，忘记密码功能不可用"}

# ---------- 物资管理 ----------
def get_material_locations(cur, material_id):
    """获取物资的所有位置及库存"""
    cur.execute("SELECT id, location, stock FROM material_locations WHERE material_id=? ORDER BY created_at", (material_id,))
    return [dict(r) for r in cur.fetchall()]

@app.get("/api/materials")
def list_materials(conn=Depends(get_db), user=Depends(get_current_user)):
    cur = conn.cursor()
    cur.execute("""
        SELECT id, name, qr_code, spec, unit, total_stock, available_stock, location, image, operator, created_at
        FROM materials ORDER BY created_at DESC
    """)
    materials = [dict(m) for m in cur.fetchall()]
    for m in materials:
        m["locations"] = get_material_locations(cur, m["id"])
    return materials

@app.get("/api/materials/search")
def search_materials(keyword: str = "", conn=Depends(get_db), user=Depends(get_current_user)):
    cur = conn.cursor()
    if keyword:
        cur.execute("""
            SELECT id, name, qr_code, spec, unit, total_stock, available_stock, location, image, operator
            FROM materials
            WHERE name LIKE ? OR qr_code LIKE ? OR spec LIKE ? OR operator LIKE ?
            ORDER BY name
        """, (f"%{keyword}%", f"%{keyword}%", f"%{keyword}%", f"%{keyword}%"))
    else:
        cur.execute("""
            SELECT id, name, qr_code, spec, unit, total_stock, available_stock, location, image, operator
            FROM materials ORDER BY name
        """)
    materials = [dict(m) for m in cur.fetchall()]
    for m in materials:
        m["locations"] = get_material_locations(cur, m["id"])
    return materials

@app.get("/api/materials/by-qr/{qr_code}")
def get_material_by_qr(qr_code: str, conn=Depends(get_db), user=Depends(get_current_user)):
    cur = conn.cursor()
    cur.execute("""
        SELECT id, name, qr_code, spec, unit, total_stock, available_stock, location, image, operator
        FROM materials WHERE qr_code = ?
    """, (qr_code,))
    material = cur.fetchone()
    if not material:
        raise HTTPException(status_code=404, detail="未找到该物资")
    return dict(material)

# ---------- 下载导入模板（必须在 /api/materials/{material_id} 前面，否则会被当成ID） ----------
@app.get("/api/materials/template")
def download_material_template(user=Depends(get_current_user)):
    if user["role"] != "admin":
        raise HTTPException(status_code=403, detail="只有管理员可以下载模板")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "物资导入模板"

    # 表头
    headers = ["物资名称", "规格", "单位", "库存数量", "存放位置"]
    header_fill = openpyxl.styles.PatternFill(start_color="43A047", end_color="43A047", fill_type="solid")
    header_font = openpyxl.styles.Font(bold=True, color="FFFFFF", size=11)
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")

    # 示例数据（第2行）
    sample_data = ["笔记本电脑", "ThinkPad X1", "台", 5, "办公室A柜"]
    for col, value in enumerate(sample_data, 1):
        cell = ws.cell(row=2, column=col, value=value)
        cell.font = openpyxl.styles.Font(color="999999", italic=True)

    # 说明（第4行开始）
    notes = [
        "【填写说明】",
        "1. 第1行为表头，请勿修改或删除",
        "2. 第2行为示例数据，导入前请删除或替换",
        "3. 从第2行开始填写您的物资数据",
        "4. 物资名称为必填项，其他列可选填",
        "5. 单位默认值为「个」，库存数量默认值为0",
        "6. 同名物资导入时会自动累加库存",
    ]
    note_font = openpyxl.styles.Font(color="E53935", size=10)
    for i, note in enumerate(notes):
        cell = ws.cell(row=4 + i, column=1, value=note)
        cell.font = note_font

    # 调整列宽
    column_widths = [20, 25, 10, 12, 20]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

    # 冻结首行
    ws.freeze_panes = "A2"

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=material_template.xlsx; filename*=UTF-8''{quote('物资导入模板.xlsx')}"}
    )

# ---------- 导出物资为 Excel（必须在动态路由前面） ----------
@app.get("/api/materials/export")
def export_materials(conn=Depends(get_db), user=Depends(get_current_user)):
    if user["role"] != "admin":
        raise HTTPException(status_code=403, detail="只有管理员可以导出物资")
    cur = conn.cursor()
    cur.execute("""
        SELECT name, qr_code, spec, unit, total_stock, available_stock, location, operator, created_at
        FROM materials ORDER BY created_at DESC
    """)
    materials = cur.fetchall()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "物资清单"

    # 表头
    headers = ["物资名称", "编号", "规格", "单位", "总库存", "可领取", "存放位置", "操作人", "创建时间"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = openpyxl.styles.Font(bold=True)
        cell.fill = openpyxl.styles.PatternFill(start_color="E8EAF6", end_color="E8EAF6", fill_type="solid")

    # 数据
    for row, m in enumerate(materials, 2):
        ws.cell(row=row, column=1, value=m["name"])
        ws.cell(row=row, column=2, value=m["qr_code"])
        ws.cell(row=row, column=3, value=m["spec"] or "")
        ws.cell(row=row, column=4, value=m["unit"] or "个")
        ws.cell(row=row, column=5, value=m["total_stock"])
        ws.cell(row=row, column=6, value=m["available_stock"])
        ws.cell(row=row, column=7, value=m["location"] or "")
        ws.cell(row=row, column=8, value=m["operator"] or "")
        ws.cell(row=row, column=9, value=m["created_at"] or "")

    # 调整列宽
    column_widths = [20, 15, 20, 8, 10, 10, 20, 12, 20]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

    # 保存到内存
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"物资清单_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=materials.xlsx; filename*=UTF-8''{quote(filename)}"}
    )

@app.get("/api/materials/{material_id}")
def get_material_detail(material_id: str, conn=Depends(get_db), user=Depends(get_current_user)):
    cur = conn.cursor()
    cur.execute("""
        SELECT id, name, qr_code, spec, unit, total_stock, available_stock, location, image, operator
        FROM materials WHERE id = ?
    """, (material_id,))
    material = cur.fetchone()
    if not material:
        raise HTTPException(status_code=404, detail="未找到该物资")
    return dict(material)

@app.post("/api/materials")
def create_material(req: MaterialCreate, conn=Depends(get_db), user=Depends(get_current_user)):
    material_id = str(uuid.uuid4())
    cur = conn.cursor()
    cur.execute("SELECT COUNT(*) FROM materials")
    count = cur.fetchone()[0]
    qr_code = f"MAT-{count + 1:06d}"
    operator = user.get("real_name", "") or user.get("username", "")
    cur.execute("""
        INSERT INTO materials (id, name, qr_code, spec, unit, total_stock, available_stock, location, image, operator)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, (material_id, req.name, qr_code, req.spec, req.unit, req.total_stock, req.total_stock, req.location, req.image, operator))
    # 创建位置记录
    if req.location:
        loc_id = str(uuid.uuid4())
        cur.execute("INSERT INTO material_locations (id, material_id, location, stock) VALUES (?, ?, ?, ?)",
                  (loc_id, material_id, req.location, req.total_stock))
    conn.commit()
    return {"id": material_id, "qr_code": qr_code, "message": "物资添加成功"}

# ---------- 公开物资查询（无需登录，用于登录页快速领取） ----------
@app.get("/api/public/materials")
def public_list_materials(conn=Depends(get_db)):
    cur = conn.cursor()
    cur.execute("""
        SELECT id, name, qr_code, spec, unit, total_stock, available_stock, location, image
        FROM materials ORDER BY name
    """)
    return [dict(m) for m in cur.fetchall()]

@app.get("/api/public/materials/search")
def public_search_materials(keyword: str = "", conn=Depends(get_db)):
    cur = conn.cursor()
    if keyword:
        cur.execute("""
            SELECT id, name, qr_code, spec, unit, total_stock, available_stock, location, image
            FROM materials
            WHERE name LIKE ? OR qr_code LIKE ? OR spec LIKE ? OR location LIKE ?
            ORDER BY name
        """, (f"%{keyword}%", f"%{keyword}%", f"%{keyword}%", f"%{keyword}%"))
    else:
        cur.execute("""
            SELECT id, name, qr_code, spec, unit, total_stock, available_stock, location, image
            FROM materials ORDER BY name
        """)
    return [dict(m) for m in cur.fetchall()]

@app.get("/api/public/materials/by-qr/{qr_code}")
def public_get_material_by_qr(qr_code: str, conn=Depends(get_db)):
    cur = conn.cursor()
    cur.execute("""
        SELECT id, name, qr_code, spec, unit, total_stock, available_stock, location, image
        FROM materials WHERE qr_code = ?
    """, (qr_code,))
    material = cur.fetchone()
    if not material:
        raise HTTPException(status_code=404, detail="未找到该物资")
    return dict(material)

# 公开查询用户未归还物资（需要账号密码验证）
@app.post("/api/public/my-borrowed")
def public_get_borrowed(req: LoginRequest, conn=Depends(get_db)):
    user = authenticate_user(req.username, req.password, conn)
    if not user:
        raise HTTPException(status_code=401, detail="账号或密码错误")
    cur = conn.cursor()
    cur.execute("""
        SELECT t.id, t.material_id, t.quantity, t.status,
               m.name as material_name, m.spec, m.unit, m.location as material_location
        FROM transactions t
        JOIN materials m ON t.material_id = m.id
        WHERE t.user_id = ? AND t.status = 'active' AND t.type = 'borrow'
        ORDER BY t.borrowed_at DESC
    """, (user["id"],))
    return [dict(r) for r in cur.fetchall()]

# ---------- 图片上传 ----------
@app.post("/api/upload")
async def upload_image(file: UploadFile = File(...), user=Depends(get_current_user)):
    ext = os.path.splitext(file.filename)[1] or ".jpg"
    filename = f"{uuid.uuid4().hex}{ext}"
    filepath = os.path.join(UPLOAD_DIR, filename)
    with open(filepath, "wb") as f:
        f.write(await file.read())
    return {"url": f"/static/uploads/{filename}", "filename": filename}

# ---------- 领取 ----------
@app.post("/api/borrow")
def borrow_material(req: BorrowRequest, conn=Depends(get_db), user=Depends(get_current_user)):
    try:
        cur = conn.cursor()
        # 从指定位置扣减库存
        if req.location:
            cur.execute("SELECT id, stock FROM material_locations WHERE material_id=? AND location=?", (req.material_id, req.location))
            loc = cur.fetchone()
            if not loc:
                raise HTTPException(status_code=400, detail=f"位置 {req.location} 不存在")
            if loc["stock"] < req.quantity:
                raise HTTPException(status_code=400, detail=f"位置 {req.location} 库存不足")
            cur.execute("UPDATE material_locations SET stock = stock - ? WHERE id=?", (req.quantity, loc["id"]))
        else:
            # 未指定位置，从第一个有库存的位置扣减
            cur.execute("SELECT id, stock, location FROM material_locations WHERE material_id=? AND stock >= ? ORDER BY stock DESC LIMIT 1", (req.material_id, req.quantity))
            loc = cur.fetchone()
            if not loc:
                raise HTTPException(status_code=400, detail="库存不足")
            cur.execute("UPDATE material_locations SET stock = stock - ? WHERE id=?", (req.quantity, loc["id"]))
            req.location = loc["location"]

        # 更新物资总可用库存
        cur.execute("""
            UPDATE materials SET available_stock = available_stock - ?
            WHERE id = ? AND available_stock >= ?
        """, (req.quantity, req.material_id, req.quantity))
        if cur.rowcount == 0:
            conn.rollback()
            raise HTTPException(status_code=400, detail="库存不足")

        cur.execute("SELECT name, available_stock FROM materials WHERE id = ?", (req.material_id,))
        m = cur.fetchone()

        # 更新操作人
        operator = user.get("real_name", "") or user.get("username", "")
        cur.execute("UPDATE materials SET operator = ? WHERE id = ?", (operator, req.material_id))
        conn.commit()

        tx_id = str(uuid.uuid4())
        cur.execute("""
            INSERT INTO transactions (id, material_id, user_id, type, quantity, status,
                                      activity_name, phone, borrow_image, borrowed_at, borrow_location)
            VALUES (?, ?, ?, 'borrow', ?, 'active', ?, ?, ?, ?, ?)
        """, (tx_id, req.material_id, user["id"], req.quantity,
              req.activity_name, req.phone, req.borrow_image, datetime.now().isoformat(), req.location))
        conn.commit()

        return {"success": True, "message": f"领取成功！{m['name']} 剩余 {m['available_stock']}",
                "material_name": m["name"], "remaining": m["available_stock"]}
    except HTTPException:
        raise
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=f"领取失败：{str(e)}")

# ---------- 归还 ----------
@app.post("/api/return")
def return_material(req: ReturnRequest, conn=Depends(get_db), user=Depends(get_current_user)):
    try:
        cur = conn.cursor()
        cur.execute("""
            SELECT COALESCE(SUM(quantity), 0) FROM transactions
            WHERE material_id = ? AND user_id = ? AND status = 'active' AND type = 'borrow'
        """, (req.material_id, user["id"]))
        total_borrowed = cur.fetchone()[0]
        if total_borrowed <= 0:
            raise HTTPException(status_code=400, detail="你没有该物资的未归还记录")
        if req.quantity > total_borrowed:
            raise HTTPException(status_code=400, detail=f"归还数量不能超过未还数量（{total_borrowed}）")

        # 归还到指定位置，不存在则新建
        return_loc = req.return_location.strip() if req.return_location else ""
        if return_loc:
            cur.execute("SELECT id, stock FROM material_locations WHERE material_id=? AND location=?", (req.material_id, return_loc))
            loc = cur.fetchone()
            if loc:
                cur.execute("UPDATE material_locations SET stock = stock + ? WHERE id=?", (req.quantity, loc["id"]))
            else:
                loc_id = str(uuid.uuid4())
                cur.execute("INSERT INTO material_locations (id, material_id, location, stock) VALUES (?, ?, ?, ?)",
                          (loc_id, req.material_id, return_loc, req.quantity))

        cur.execute("UPDATE materials SET available_stock = available_stock + ? WHERE id = ?",
                    (req.quantity, req.material_id))
        conn.commit()
        cur.execute("SELECT name, available_stock FROM materials WHERE id = ?", (req.material_id,))
        m = cur.fetchone()

        # 更新操作人
        operator = user.get("real_name", "") or user.get("username", "")
        cur.execute("UPDATE materials SET operator = ? WHERE id = ?", (operator, req.material_id))
        conn.commit()

        cur.execute("""
            UPDATE transactions SET status = 'completed', returned_at = ?,
                   return_time = ?, return_location = ?, return_image = ?
            WHERE id IN (
                SELECT id FROM transactions
                WHERE material_id = ? AND user_id = ? AND status = 'active' AND type = 'borrow'
                ORDER BY borrowed_at ASC LIMIT 1
            )
        """, (datetime.now().isoformat(), req.return_time, req.return_location, req.return_image,
              req.material_id, user["id"]))
        conn.commit()

        return {"success": True, "message": f"归还成功！{m['name']} 当前可领 {m['available_stock']}",
                "material_name": m["name"], "remaining": m["available_stock"]}
    except HTTPException:
        raise
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=f"归还失败：{str(e)}")

# ---------- 公开领取（登录信息+领取信息合并提交，无需预先登录） ----------
@app.post("/api/public/borrow")
def public_borrow(req: PublicBorrowRequest, conn=Depends(get_db)):
    # 先验证登录
    user = authenticate_user(req.username, req.password, conn)
    if not user:
        raise HTTPException(status_code=401, detail="账号或密码错误")
    try:
        cur = conn.cursor()
        # 从指定位置扣减库存
        if req.location:
            cur.execute("SELECT id, stock FROM material_locations WHERE material_id=? AND location=?", (req.material_id, req.location))
            loc = cur.fetchone()
            if not loc:
                raise HTTPException(status_code=400, detail=f"位置 {req.location} 不存在")
            if loc["stock"] < req.quantity:
                raise HTTPException(status_code=400, detail=f"位置 {req.location} 库存不足")
            cur.execute("UPDATE material_locations SET stock = stock - ? WHERE id=?", (req.quantity, loc["id"]))
        else:
            cur.execute("SELECT id, stock, location FROM material_locations WHERE material_id=? AND stock >= ? ORDER BY stock DESC LIMIT 1", (req.material_id, req.quantity))
            loc = cur.fetchone()
            if not loc:
                raise HTTPException(status_code=400, detail="库存不足")
            cur.execute("UPDATE material_locations SET stock = stock - ? WHERE id=?", (req.quantity, loc["id"]))
            req.location = loc["location"]

        cur.execute("""
            UPDATE materials SET available_stock = available_stock - ?
            WHERE id = ? AND available_stock >= ?
        """, (req.quantity, req.material_id, req.quantity))
        if cur.rowcount == 0:
            conn.rollback()
            raise HTTPException(status_code=400, detail="库存不足")

        cur.execute("SELECT name, available_stock FROM materials WHERE id = ?", (req.material_id,))
        m = cur.fetchone()

        operator = user.get("real_name", "") or user.get("username", "")
        cur.execute("UPDATE materials SET operator = ? WHERE id = ?", (operator, req.material_id))
        conn.commit()

        tx_id = str(uuid.uuid4())
        cur.execute("""
            INSERT INTO transactions (id, material_id, user_id, type, quantity, status,
                                      activity_name, phone, borrow_image, borrowed_at, borrow_location)
            VALUES (?, ?, ?, 'borrow', ?, 'active', ?, ?, ?, ?, ?)
        """, (tx_id, req.material_id, user["id"], req.quantity,
              req.activity_name, req.phone, req.borrow_image, datetime.now().isoformat(), req.location))
        conn.commit()

        # 自动生成token返回，方便后续操作
        token = create_access_token({"sub": user["username"], "id": user["id"]})
        return {"success": True, "message": f"领取成功！{m['name']} 剩余 {m['available_stock']}",
                "material_name": m["name"], "remaining": m["available_stock"],
                "token": token, "user": {"id": user["id"], "username": user["username"],
                                          "real_name": user["real_name"], "role": user["role"], "phone": user["phone"]}}
    except HTTPException:
        raise
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=f"领取失败：{str(e)}")

# ---------- 公开归还（登录信息+归还信息合并提交，无需预先登录） ----------
@app.post("/api/public/return")
def public_return(req: PublicReturnRequest, conn=Depends(get_db)):
    # 先验证登录
    user = authenticate_user(req.username, req.password, conn)
    if not user:
        raise HTTPException(status_code=401, detail="账号或密码错误")
    try:
        cur = conn.cursor()
        cur.execute("""
            SELECT COALESCE(SUM(quantity), 0) FROM transactions
            WHERE material_id = ? AND user_id = ? AND status = 'active' AND type = 'borrow'
        """, (req.material_id, user["id"]))
        total_borrowed = cur.fetchone()[0]
        if total_borrowed <= 0:
            raise HTTPException(status_code=400, detail="你没有该物资的未归还记录")
        if req.quantity > total_borrowed:
            raise HTTPException(status_code=400, detail=f"归还数量不能超过未还数量（{total_borrowed}）")

        # 归还到指定位置，不存在则新建
        return_loc = req.return_location.strip() if req.return_location else ""
        if return_loc:
            cur.execute("SELECT id, stock FROM material_locations WHERE material_id=? AND location=?", (req.material_id, return_loc))
            loc = cur.fetchone()
            if loc:
                cur.execute("UPDATE material_locations SET stock = stock + ? WHERE id=?", (req.quantity, loc["id"]))
            else:
                loc_id = str(uuid.uuid4())
                cur.execute("INSERT INTO material_locations (id, material_id, location, stock) VALUES (?, ?, ?, ?)",
                          (loc_id, req.material_id, return_loc, req.quantity))

        cur.execute("UPDATE materials SET available_stock = available_stock + ? WHERE id = ?",
                    (req.quantity, req.material_id))
        conn.commit()
        cur.execute("SELECT name, available_stock FROM materials WHERE id = ?", (req.material_id,))
        m = cur.fetchone()

        operator = user.get("real_name", "") or user.get("username", "")
        cur.execute("UPDATE materials SET operator = ? WHERE id = ?", (operator, req.material_id))
        conn.commit()

        cur.execute("""
            UPDATE transactions SET status = 'completed', returned_at = ?,
                   return_time = ?, return_location = ?, return_image = ?
            WHERE id IN (
                SELECT id FROM transactions
                WHERE material_id = ? AND user_id = ? AND status = 'active' AND type = 'borrow'
                ORDER BY borrowed_at ASC LIMIT 1
            )
        """, (datetime.now().isoformat(), req.return_time, req.return_location, req.return_image,
              req.material_id, user["id"]))
        conn.commit()

        token = create_access_token({"sub": user["username"], "id": user["id"]})
        return {"success": True, "message": f"归还成功！{m['name']} 当前可领 {m['available_stock']}",
                "material_name": m["name"], "remaining": m["available_stock"],
                "token": token, "user": {"id": user["id"], "username": user["username"],
                                          "real_name": user["real_name"], "role": user["role"], "phone": user["phone"]}}
    except HTTPException:
        raise
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=f"归还失败：{str(e)}")

# ---------- 我的记录 ----------
@app.get("/api/my-transactions")
def my_transactions(conn=Depends(get_db), user=Depends(get_current_user)):
    cur = conn.cursor()
    cur.execute("""
        SELECT t.id, t.material_id, t.quantity, t.status, t.activity_name, t.phone, t.borrow_image,
               t.return_time, t.return_location, t.return_image, t.borrowed_at, t.returned_at,
               m.name as material_name, m.spec, m.unit, m.image as material_image, m.location as material_location
        FROM transactions t
        JOIN materials m ON t.material_id = m.id
        WHERE t.user_id = ?
        ORDER BY t.borrowed_at DESC LIMIT 200
    """, (user["id"],))
    return [dict(r) for r in cur.fetchall()]

@app.get("/api/materials/{material_id}/my-borrowed")
def my_borrowed_of_material(material_id: str, conn=Depends(get_db), user=Depends(get_current_user)):
    cur = conn.cursor()
    cur.execute("""
        SELECT COALESCE(SUM(quantity), 0) FROM transactions
        WHERE material_id = ? AND user_id = ? AND status = 'active' AND type = 'borrow'
    """, (material_id, user["id"]))
    return {"borrowed_quantity": cur.fetchone()[0]}

# ---------- 管理员：所有记录 ----------
@app.get("/api/transactions")
def all_transactions(conn=Depends(get_db), user=Depends(get_current_user)):
    if user["role"] != "admin":
        raise HTTPException(status_code=403, detail="只有管理员可以查看所有记录")
    cur = conn.cursor()
    cur.execute("""
        SELECT t.id, t.quantity, t.status, t.activity_name, t.phone, t.borrow_image,
               t.return_time, t.return_location, t.return_image, t.borrowed_at, t.returned_at,
               m.name as material_name, m.spec, m.unit,
               u.real_name as user_name, u.role as user_role
        FROM transactions t
        JOIN materials m ON t.material_id = m.id
        JOIN users u ON t.user_id = u.id
        ORDER BY t.borrowed_at DESC LIMIT 500
    """)
    return [dict(r) for r in cur.fetchall()]

# ---------- 编辑物资 ----------
@app.put("/api/materials/{material_id}")
def update_material(material_id: str, req: MaterialUpdate, conn=Depends(get_db), user=Depends(get_current_user)):
    if user["role"] != "admin":
        raise HTTPException(status_code=403, detail="只有管理员可以编辑物资")
    cur = conn.cursor()
    cur.execute("SELECT * FROM materials WHERE id=?", (material_id,))
    material = cur.fetchone()
    if not material:
        raise HTTPException(status_code=404, detail="物资不存在")
    
    updates = []
    params = []
    
    if req.name is not None:
        if not req.name.strip():
            raise HTTPException(status_code=400, detail="物资名称不能为空")
        updates.append("name=?")
        params.append(req.name.strip())
    
    if req.spec is not None:
        updates.append("spec=?")
        params.append(req.spec)
    
    if req.unit is not None:
        updates.append("unit=?")
        params.append(req.unit)
    
    if req.location is not None:
        updates.append("location=?")
        params.append(req.location)
    
    if req.image is not None:
        updates.append("image=?")
        params.append(req.image)
    
    if req.total_stock is not None:
        if req.total_stock < 0:
            raise HTTPException(status_code=400, detail="库存数量不能为负数")
        old_total = material["total_stock"]
        old_available = material["available_stock"]
        diff = req.total_stock - old_total
        new_available = old_available + diff
        if new_available < 0:
            raise HTTPException(status_code=400, detail=f"库存调整后可用库存将为负数（当前可用{old_available}，最多可减少到{old_available}）")
        updates.append("total_stock=?")
        params.append(req.total_stock)
        updates.append("available_stock=?")
        params.append(new_available)
    
    if updates:
        params.append(material_id)
        cur.execute(f"UPDATE materials SET {', '.join(updates)} WHERE id=?", params)
        conn.commit()
    
    return {"success": True, "message": "物资更新成功"}

# ---------- 删除物资 ----------
@app.delete("/api/materials/{material_id}")
def delete_material(material_id: str, conn=Depends(get_db), user=Depends(get_current_user)):
    if user["role"] != "admin":
        raise HTTPException(status_code=403, detail="只有管理员可以删除物资")
    cur = conn.cursor()
    # 检查是否有未归还的领用
    cur.execute("SELECT COALESCE(SUM(quantity),0) FROM transactions WHERE material_id=? AND status='active'", (material_id,))
    if cur.fetchone()[0] > 0:
        raise HTTPException(status_code=400, detail="该物资还有未归还的领用记录，无法删除")
    cur.execute("DELETE FROM materials WHERE id=?", (material_id,))
    cur.execute("DELETE FROM transactions WHERE material_id=?", (material_id,))
    conn.commit()
    return {"success": True, "message": "物资已删除"}

# ---------- 批量删除物资 ----------
@app.post("/api/materials/batch-delete")
def batch_delete_materials(req: dict, conn=Depends(get_db), user=Depends(get_current_user)):
    if user["role"] != "admin":
        raise HTTPException(status_code=403, detail="只有管理员可以删除物资")
    ids = req.get("ids", [])
    if not ids:
        raise HTTPException(status_code=400, detail="请选择要删除的物资")
    cur = conn.cursor()
    deleted = 0
    skipped = []
    for mid in ids:
        cur.execute("SELECT COALESCE(SUM(quantity),0) FROM transactions WHERE material_id=? AND status='active'", (mid,))
        if cur.fetchone()[0] > 0:
            skipped.append(mid)
            continue
        cur.execute("DELETE FROM materials WHERE id=?", (mid,))
        cur.execute("DELETE FROM transactions WHERE material_id=?", (mid,))
        deleted += 1
    conn.commit()
    msg = f"成功删除 {deleted} 条"
    if skipped:
        msg += f"，{len(skipped)} 条因有未归还记录跳过"
    return {"success": True, "message": msg, "deleted": deleted, "skipped": len(skipped)}

# ---------- 更新物资图片 ----------
@app.post("/api/materials/{material_id}/image")
def update_material_image(material_id: str, file: UploadFile = File(...), conn=Depends(get_db), user=Depends(get_current_user)):
    if user["role"] != "admin":
        raise HTTPException(status_code=403, detail="只有管理员可以修改物资图片")
    ext = os.path.splitext(file.filename)[1] or ".jpg"
    filename = f"{uuid.uuid4().hex}{ext}"
    filepath = os.path.join(UPLOAD_DIR, filename)
    with open(filepath, "wb") as f:
        f.write(file.file.read())
    cur = conn.cursor()
    cur.execute("UPDATE materials SET image=? WHERE id=?", (f"/static/uploads/{filename}", material_id))
    conn.commit()
    return {"success": True, "message": "图片更新成功", "image": f"/static/uploads/{filename}"}

# ---------- 数据备份 ----------
@app.get("/api/backup")
def backup_database(user=Depends(get_current_user)):
    if user["role"] != "admin":
        raise HTTPException(status_code=403, detail="只有管理员可以备份数据")
    if not os.path.exists(DATABASE_FILE):
        raise HTTPException(status_code=404, detail="数据库文件不存在")
    from datetime import datetime as dt
    filename = f"物资管理系统_备份_{dt.now().strftime('%Y%m%d_%H%M%S')}.db"
    with open(DATABASE_FILE, "rb") as f:
        content = f.read()
    return Response(
        content=content,
        media_type="application/octet-stream",
        headers={"Content-Disposition": f"attachment; filename=backup.db; filename*=UTF-8''{quote(filename)}"}
    )

# ---------- 数据恢复 ----------
@app.post("/api/restore")
def restore_database(file: UploadFile = File(...), user=Depends(get_current_user)):
    if user["role"] != "admin":
        raise HTTPException(status_code=403, detail="只有管理员可以恢复数据")
    if not file.filename.endswith('.db'):
        raise HTTPException(status_code=400, detail="请上传 .db 格式的备份文件")
    try:
        content = file.file.read()
        # 验证是否是有效的 SQLite 数据库文件（前16字节是 SQLite header）
        if not content.startswith(b"SQLite format 3"):
            raise HTTPException(status_code=400, detail="文件不是有效的 SQLite 数据库备份")
        # 先备份当前数据库
        backup_path = DATABASE_FILE + ".bak_" + str(int(time.time()))
        if os.path.exists(DATABASE_FILE):
            import shutil
            shutil.copy2(DATABASE_FILE, backup_path)
        # 写入新数据库
        with open(DATABASE_FILE, "wb") as f:
            f.write(content)
        return {"success": True, "message": "数据恢复成功，请刷新页面。当前数据库已自动备份为 " + os.path.basename(backup_path)}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"恢复失败: {str(e)}")

# ---------- 从 Excel 导入物资 ----------
@app.post("/api/materials/import")
def import_materials(file: UploadFile = File(...), conn=Depends(get_db), user=Depends(get_current_user)):
    if user["role"] != "admin":
        raise HTTPException(status_code=403, detail="只有管理员可以导入物资")

    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="请上传 Excel 文件（.xlsx 格式）")

    try:
        contents = file.file.read()
        wb = openpyxl.load_workbook(BytesIO(contents))
        ws = wb.active

        # 读取表头，确认列顺序
        headers = [cell.value for cell in ws[1]]
        # 找到各列的索引（支持中文表头）
        col_map = {}
        for idx, h in enumerate(headers):
            if h and ("名称" in str(h) or "物资" in str(h)):
                col_map["name"] = idx
            elif h and "规格" in str(h):
                col_map["spec"] = idx
            elif h and "单位" in str(h):
                col_map["unit"] = idx
            elif h and ("库存" in str(h) or "数量" in str(h)):
                col_map["stock"] = idx
            elif h and ("位置" in str(h) or "存放" in str(h)):
                col_map["location"] = idx

        if "name" not in col_map:
            raise HTTPException(status_code=400, detail="Excel 表头必须包含「物资名称」列")

        cur = conn.cursor()
        operator = user.get("real_name", "") or user.get("username", "")
        added_count = 0
        updated_count = 0
        error_rows = []

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
            try:
                name = row[col_map["name"]]
                if not name or str(name).strip() == "":
                    continue
                name = str(name).strip()
                spec = str(row[col_map.get("spec", -1)]).strip() if col_map.get("spec") is not None and row[col_map["spec"]] else ""
                unit = str(row[col_map.get("unit", -1)]).strip() if col_map.get("unit") is not None and row[col_map["unit"]] else "个"
                stock = int(row[col_map.get("stock", -1)]) if col_map.get("stock") is not None and row[col_map["stock"]] else 0
                location = str(row[col_map.get("location", -1)]).strip() if col_map.get("location") is not None and row[col_map["location"]] else ""

                # 检查物资是否已存在
                cur.execute("SELECT id, total_stock, available_stock FROM materials WHERE name = ?", (name,))
                existing = cur.fetchone()

                if existing:
                    # 更新已有物资：累加库存
                    new_total = existing["total_stock"] + stock
                    new_available = existing["available_stock"] + stock
                    cur.execute("""
                        UPDATE materials SET spec=?, unit=?, total_stock=?, available_stock=?, location=?, operator=?
                        WHERE id=?
                    """, (spec, unit, new_total, new_available, location, operator, existing["id"]))
                    updated_count += 1
                else:
                    # 新增物资
                    material_id = str(uuid.uuid4())
                    cur.execute("SELECT COUNT(*) FROM materials")
                    count = cur.fetchone()[0]
                    qr_code = f"MAT-{count + 1:06d}"
                    cur.execute("""
                        INSERT INTO materials (id, name, qr_code, spec, unit, total_stock, available_stock, location, image, operator)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """, (material_id, name, qr_code, spec, unit, stock, stock, location, "", operator))
                    added_count += 1
            except Exception as e:
                error_rows.append({"row": row_idx, "error": str(e)})

        conn.commit()
        return {
            "success": True,
            "message": f"导入完成：新增 {added_count} 条，更新 {updated_count} 条",
            "added": added_count,
            "updated": updated_count,
            "errors": error_rows
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"导入失败：{str(e)}")

# ---------- 删除记录 ----------
@app.delete("/api/transactions/{tx_id}")
def delete_transaction(tx_id: str, conn=Depends(get_db), user=Depends(get_current_user)):
    if user["role"] != "admin":
        raise HTTPException(status_code=403, detail="只有管理员可以删除记录")
    cur = conn.cursor()
    cur.execute("DELETE FROM transactions WHERE id=?", (tx_id,))
    conn.commit()
    return {"success": True, "message": "记录已删除"}

# ---------- 数据大屏统计 ----------
@app.get("/api/stats/overview")
def stats_overview(conn=Depends(get_db), user=Depends(get_current_user)):
    cur = conn.cursor()

    # 物资总数
    cur.execute("SELECT COUNT(*) as cnt, COALESCE(SUM(total_stock),0) as total, COALESCE(SUM(available_stock),0) as available FROM materials")
    m = cur.fetchone()
    material_count = m["cnt"]
    total_stock = m["total"]
    available_stock = m["available"]
    borrowed_stock = total_stock - available_stock

    # 领取总数
    cur.execute("SELECT COUNT(*) as cnt, COALESCE(SUM(quantity),0) as total FROM transactions WHERE type='borrow'")
    b = cur.fetchone()
    borrow_count = b["cnt"]
    borrow_total = b["total"]

    # 归还总数
    cur.execute("SELECT COUNT(*) as cnt, COALESCE(SUM(quantity),0) as total FROM transactions WHERE type='borrow' AND status='completed'")
    r = cur.fetchone()
    return_count = r["cnt"]
    return_total = r["total"]

    # 存放位置分布
    cur.execute("""
        SELECT COALESCE(NULLIF(location,''),'未分类') as loc, COUNT(*) as cnt
        FROM materials GROUP BY location ORDER BY cnt DESC
    """)
    location_dist = [dict(row) for row in cur.fetchall()]

    # 库存状态分布（充足/偏低/为零）
    cur.execute("SELECT COUNT(*) FROM materials WHERE available_stock > 5")
    stock_ok = cur.fetchone()[0]
    cur.execute("SELECT COUNT(*) FROM materials WHERE available_stock > 0 AND available_stock <= 5")
    stock_low = cur.fetchone()[0]
    cur.execute("SELECT COUNT(*) FROM materials WHERE available_stock = 0")
    stock_out = cur.fetchone()[0]

    # 最近7天领取趋势
    trend = []
    for i in range(6, -1, -1):
        date = (datetime.now() - timedelta(days=i)).strftime("%Y-%m-%d")
        cur.execute("""
            SELECT COALESCE(SUM(quantity),0) as total FROM transactions
            WHERE type='borrow' AND DATE(borrowed_at) = ?
        """, (date,))
        trend.append({"date": date[5:], "count": cur.fetchone()[0]})

    return {
        "material_count": material_count,
        "total_stock": total_stock,
        "available_stock": available_stock,
        "borrowed_stock": borrowed_stock,
        "borrow_count": borrow_count,
        "borrow_total": borrow_total,
        "return_count": return_count,
        "return_total": return_total,
        "location_dist": location_dist,
        "stock_status": {"ok": stock_ok, "low": stock_low, "out": stock_out},
        "borrow_trend": trend,
    }

# ==================== 静态文件 ====================
app.mount("/static", StaticFiles(directory="static"), name="static")

@app.get("/")
def index():
    return FileResponse("static/index.html")

@app.get("/login")
def web_login():
    return FileResponse("static/login.html")

@app.get("/dashboard")
def dashboard():
    return FileResponse("static/dashboard.html")

@app.get("/m/")
def mobile_index():
    return FileResponse("static/m/index.html")

@app.get("/m/admin")
def mobile_admin():
    return FileResponse("static/m/admin.html")

@app.get("/m/login")
def mobile_login():
    return FileResponse("static/m/login.html")

@app.get("/m/scan.html")
def mobile_scan():
    return FileResponse("static/m/scan.html")

# ==================== 启动 ====================
if __name__ == "__main__":
    import uvicorn
    port = int(os.environ.get("PORT", 8000))
    print("=" * 50)
    print("  城治学生会物资管理系统（升级版）启动中...")
    print(f"  服务端口: {port}")
    print("  API 文档: /docs")
    print("=" * 50)
    uvicorn.run(app, host="0.0.0.0", port=port)
